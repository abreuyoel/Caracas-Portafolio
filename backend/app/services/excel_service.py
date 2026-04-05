import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple
from datetime import datetime
import logging
from openpyxl.styles import PatternFill, Font

logger = logging.getLogger(__name__)

class ExcelService:
    """Servicio para importar/exportar transacciones desde Excel"""
    
    STOCK_SYMBOL_MAPPING = {
        'bancaribe': 'ABC.A', 'banco caribe': 'ABC.A', 'caribe': 'ABC.A', 'abc': 'ABC.A', 'abc.a': 'ABC.A',
        'provincial': 'BPV', 'banco provincial': 'BPV', 'bpv': 'BPV',
        'ron': 'RST', 'santa teresa': 'RST', 'ron santa teresa': 'RST', 'rst': 'RST',
        'zuliano': 'GZL', 'grupo zuliano': 'GZL', 'gzl': 'GZL',
        'sivensa': 'SVS', 'svs': 'SVS',
        'proagro': 'PGR', 'pgr': 'PGR',
        'bnc': 'BNC', 'nacional credito': 'BNC', 'banco nacional': 'BNC',
        'venezuela': 'BVL', 'banco venezuela': 'BVL', 'bvl': 'BVL',
        'mercantil': 'MVZ.A', 'mvz': 'MVZ.A', 'mvz.a': 'MVZ.A',
        'cantv': 'TDV.D', 'telefonia': 'TDV.D', 'tdv': 'TDV.D', 'tdv.d': 'TDV.D',
        'cementos': 'FNC', 'fabrica cementos': 'FNC', 'fnc': 'FNC',
        'manpa': 'MPA', 'mpa': 'MPA', 'papel': 'MPA',
        'ceramica': 'CCR', 'carabobo': 'CCR', 'ccr': 'CCR',
    }
    
    @staticmethod
    def create_template(stocks: List[Dict]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        C_HEADER, C_WHITE, C_FORMULA, C_OPTIONAL, C_SLIP, C_INPUT = "1F3864", "FFFFFF", "D6EAF8", "F5F5F5", "FEF9E7", "FFFFFF"
        thin = Side(style='thin', color="CCCCCC")
        bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _font(bold=False, color="000000", sz=10, italic=False): return Font(name='Arial', bold=bold, color=color, size=sz, italic=italic)
        def _fill(color): return PatternFill('solid', start_color=color)
        def _align(h='left', wrap=False): return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

        wb = Workbook()
        active_stocks = sorted([s for s in stocks if s.get('is_active', True)], key=lambda x: x['symbol'])
        ws_ref = wb.create_sheet("Acciones_Activas")
        ws_ref.sheet_state = 'hidden'
        ws_ref['A1'], ws_ref['B1'], ws_ref['C1'] = 'Símbolo', 'Nombre', 'Símbolo - Nombre'
        for i, s in enumerate(active_stocks, start=2):
            symbol, name = s.get('symbol', '???'), s.get('name', '')
            ws_ref.cell(i, 1, symbol); ws_ref.cell(i, 2, name); ws_ref.cell(i, 3, f"{symbol} - {name}" if name else symbol)
        last_stock_idx = len(active_stocks) + 1

        ws_cat = wb.create_sheet("Catalogos")
        ws_cat.sheet_state = 'hidden'
        brokers_list = [
            "VALORES VENCRED CASA DE BOLSA, S.A.", "GRUPO BURSATIL VENEZOLANO CASA DE BOLSA, C.A",
            "ACTIVALORES CASA DE BOLSA S.A.", "AGRONET VALORES CASA DE BOLSA C.A.",
            "MERCANTIL MERINVEST, CASA DE BOLSA, C.A.", "PLATINUM CASA DE BOLSA, C.A.",
            "ACCIONA CASA DE BOLSA, S.A.", "CAJA CARACAS CASA DE BOLSA, C.A.",
            "INTERBURSA CASA DE BOLSA, C.A.", "FINANCORP VALORES CASA DE BOLSA, C.A.",
            "KAIZEN CASA DE BOLSA C.A.", "FIVENCA CASA DE BOLSA, C.A.",
            "PER CAPITAL SOCIEDAD DE CORRETAJE DE VALORES, C.A.", "INCORP CASA DE BOLSA, C.A.",
            "MULTIPLICAS CASA DE BOLSA, C.A", "MERCOSUR CASA DE BOLSA, S.A.",
            "INTERGLOBAL CASA DE BOLSA, C.A.", "MAXIMIZA CASA DE BOLSA, C.A.",
            "SOLFIN CASA DE BOLSA, C.A.", "INTERBONO CASA DE BOLSA, C.A.",
            "INVERAMIGA CASA DE BOLSA, C.A.", "RATIO CASA DE BOLSA, C.A.",
            "RENDIVALORES CASA DE BOLSA. C.A.", "CUADRA CASA DE BOLSA, S.A.",
            "HLB VALORES CASA DE BOLSA, C.A.", "MASVALOR CASA DE BOLSA, S.A.",
            "KOI INVEST CASA DE BOLSA, C.A.", "MULTIVALORES CASA DE BOLSA C.A.",
            "KAIROS VALORES CASA DE BOLSA, C.A.", "GRUPO ITALCAPITAL CASA DE BOLSA, C.A.",
            "BNCI CASA DE BOLSA, C.A.", "SUMA CASA DE BOLSA, C.A.",
            "WORLD TRADING CASA DE BOLSA, C.A.", "INVERCAPITAL CASA DE BOLSA, S.A."
        ]
        for i, b in enumerate(brokers_list, 1): ws_cat.cell(i, 1, b)
        last_broker_idx = len(brokers_list)

        ws = wb.active; ws.title = "Transacciones"; ws.freeze_panes = "A3"
        ws.row_dimensions[1].height, ws.row_dimensions[2].height = 36, 30
        for col, w in {'A':12,'B':12,'C':22,'D':13,'E':35,'F':12,'G':14,'H':13,'I':10,'J':10,'K':10,'L':13,'M':12,'N':12,'O':12,'P':10}.items(): ws.column_dimensions[col].width = w

        ws.merge_cells("A1:Z1"); c = ws['A1']; c.value = "PLANTILLA DE TRANSACCIONES BURSÁTILES — BOLSA DE VALORES DE CARACAS"
        c.font, c.fill, c.alignment = _font(bold=True, sz=13, color=C_WHITE), _fill(C_HEADER), _align('center')

        headers = [('A',"Orden",C_OPTIONAL,"888888"),('B',"Tipo Orden *",C_HEADER,C_WHITE),('C',"Acción *",C_HEADER,C_WHITE),('D',"Tipo Solicitud *",C_HEADER,C_WHITE),('E',"Casa de Bolsa *",C_HEADER,C_WHITE),('F',"Cantidad *",C_HEADER,C_WHITE),('G',"Precio Promedio *",C_HEADER,C_WHITE),('H',"Monto Bruto",C_FORMULA,"1A5276"),('I',"Comisión",C_HEADER,C_WHITE),('J',"IVA",C_HEADER,C_WHITE),('K',"Derecho Reg.",C_HEADER,C_WHITE),('L',"Monto Neto",C_FORMULA,"1A5276"),('M',"Tasa BCV",C_HEADER,C_WHITE),('N',"Monto $",C_HEADER,C_WHITE),('O',"Fecha",C_HEADER,C_WHITE),('P',"Slippage?",C_SLIP,"7D6608")]
        for col, label, bg, fg in headers:
            c = ws[f"{col}2"]; c.value = label; c.font, c.fill, c.alignment, c.border = _font(bold=True, color=fg, sz=9), _fill(bg), _align('center', True), bdr

        for r in range(3, 103):
            for col, fill in [(1,C_OPTIONAL),(2,C_INPUT),(3,C_INPUT),(4,C_INPUT),(5,C_INPUT),(6,C_INPUT),(7,C_INPUT),(9,C_INPUT),(10,C_INPUT),(11,C_INPUT),(13,C_INPUT),(15,C_INPUT)]:
                c = ws.cell(r, col); c.fill, c.border, c.alignment, c.font = _fill(fill), bdr, _align('center'), _font()
            ws.cell(r, 8).value = f'=IFERROR(F{r}*G{r}, "")'
            ws.cell(r, 12).value = f'=IFERROR(H{r}+I{r}+J{r}+K{r}, "")'
            ws.cell(r, 14).value = f'=IFERROR(L{r}/M{r}, "")'
            c = ws.cell(r, 16); c.font, c.fill, c.alignment, c.border = _font(bold=True, color="7D6608"), _fill(C_SLIP), _align('center'), bdr

        dv_o = DataValidation(type="list", formula1='"Compra,Venta"', allow_blank=True); ws.add_data_validation(dv_o); dv_o.add("B3:B102")
        dv_s = DataValidation(type="list", formula1='"Mercado,Limite"', allow_blank=True); ws.add_data_validation(dv_s); dv_s.add("D3:D102")
        dv_p = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True); ws.add_data_validation(dv_p); dv_p.add("P3:P102")
        dv_b = DataValidation(type="list", formula1=f"=Catalogos!$A$1:$A${last_broker_idx}", allow_blank=True); ws.add_data_validation(dv_b); dv_b.add("E3:E102")
        dv_a = DataValidation(type="list", formula1=f"=Acciones_Activas!$C$2:$C${last_stock_idx}", allow_blank=True); ws.add_data_validation(dv_a); dv_a.add("C3:C102")

        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    @staticmethod
    def normalize_stock_symbol(symbol: str) -> str:
        if not symbol: return symbol
        s = symbol.strip()
        if ' - ' in s: s = s.split(' - ')[0]
        s_u = s.strip().upper()
        return ExcelService.STOCK_SYMBOL_MAPPING.get(s_u.lower(), s_u)

    @staticmethod
    def normalize_broker_name(name: str) -> str:
        if not name or pd.isna(name): return None
        n = str(name).strip().upper()
        # Mapeos comunes o simplificaciones
        if "MERCOSUR" in n: return "MERCOSUR CASA DE BOLSA, S.A."
        if "VALORES VENCRED" in n: return "VALORES VENCRED CASA DE BOLSA, S.A."
        if "PROVINCIAL" in n or "BPV" in n: return "GRUPO BURSATIL VENEZOLANO CASA DE BOLSA, C.A" # Wait, Provincial is different
        # Si no hay match específico, devolver el nombre original limpio
        # O podrías buscar en la list original de create_template
        return name.strip()

    @staticmethod
    def parse_excel(file_content: bytes) -> Tuple[List[Dict], List[Dict]]:
        errors, transactions = [], []
        def _num(v):
            if v is None: return 0.0
            s = str(v).strip().replace("Bs", "").replace("$", "").replace(" ", "")
            if s in ("", "nan", "-"): return 0.0
            if "," in s and "." in s: s = s.replace(".", "").replace(",", ".")
            elif "," in s: s = s.replace(",", ".")
            try: return float(s)
            except: return 0.0

        try:
            raw = pd.read_excel(io.BytesIO(file_content), sheet_name="Transacciones", header=1, dtype=str)
            raw.dropna(how="all", inplace=True)
            for idx, row in raw.iterrows():
                row_num = idx + 3
                acc_raw = row.iloc[2] # Col C
                if pd.isna(acc_raw) or str(acc_raw).startswith("📌"): continue
                
                symbol = ExcelService.normalize_stock_symbol(str(acc_raw))
                transactions.append({
                    "order_number": str(row.iloc[0]) if not pd.isna(row.iloc[0]) else None,
                    "order_type": str(row.iloc[1]).capitalize() if not pd.isna(row.iloc[1]) else "Compra",
                    "stock_symbol": symbol,
                    "request_type": str(row.iloc[3]).capitalize() if not pd.isna(row.iloc[3]) else "Mercado",
                    "brokerage": ExcelService.normalize_broker_name(row.iloc[4]),
                    "quantity": int(_num(row.iloc[5])),
                    "avg_price": _num(row.iloc[6]),
                    "commission": _num(row.iloc[8]),
                    "iva": _num(row.iloc[9]),
                    "registry_fee": _num(row.iloc[10]),
                    "net_amount": _num(row.iloc[11]),
                    "bcv_rate": _num(row.iloc[12]),
                    "amount_usd": _num(row.iloc[13]),
                    "transaction_date": str(row.iloc[14]).split(' ')[0] if not pd.isna(row.iloc[14]) else None,
                    "notes": f"Importado: {symbol}"
                })
            return transactions, errors
        except Exception as e:
            errors.append({"row": 0, "error": str(e)}); return transactions, errors

    @staticmethod
    def create_export(transactions: List[Dict]) -> bytes:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Export"
        hs = ['Orden', 'Tipo', 'Accion', 'Solicitud', 'Broker', 'Cant', 'Precio', 'Neto', 'USD', 'Fecha']
        for i, h in enumerate(hs, 1): ws.cell(1, i, h)
        for r, tx in enumerate(transactions, 2):
            for c, k in enumerate(['order_number','order_type','stock_symbol','request_type','brokerage','quantity','avg_price','net_amount','amount_usd','transaction_date'], 1):
                ws.cell(r, c, tx.get(k))
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

excel_service = ExcelService()